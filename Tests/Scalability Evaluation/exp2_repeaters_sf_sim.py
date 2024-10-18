import sys
import openpyxl
import argparse
import matplotlib.pyplot as plt
sys.path.append('../../')
import loraMeshSimulator as sim

def append_values_to_excel(file_path, sheet_name, values):
    # Load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the specified sheet or create a new one if it doesn't exist
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    # Append values to a new row
    sheet.append(values)

    # Save the changes to the workbook
    workbook.save(file_path)

def main(load, density, no_of_repeaters, repeater_delay_multiplier):
    #-------------------------------------------------------------------------------------
    # Simulation config
    #-------------------------------------------------------------------------------------
    node = sim.node
    nodes = sim.nodes
    env =sim.env
    maxDist =sim.maxDist

    sim.avgSendTime = 60000 # 1 mins
    sim.repeatDelayMultiplier = repeater_delay_multiplier
    sim.graphics = 0
    sim.realtime_graphics = 0
    sim.debug = 0

    sim.totalSimPackets = 5000

    repeaters =[]
    enddevices = []

    for i in range(no_of_repeaters):
        repeaters.append(node(env, (1+float(i)/density)*maxDist*0.99, 1*maxDist*1, "rp"))
        for j in range(load):
            enddevices.append(node(env, (1+float(i)/density)*maxDist*0.99, 2*maxDist, "ed"))

    gw = sim.node(env, (1+float(no_of_repeaters)/density)*maxDist*0.99, 1*maxDist*1, "gw")

    predicted_DER_List = [0.55, 0.53, 0.52, 0.52, 0.52, 0.52, 0.52, 0.52, 0.52, 0.52, 0.54, 0.56, 0.60, 0.61, 0.63, 0.6, 0.66, 0.68, 0.68, 0.70, 0.72, 0.74, 0.74, 0.74, 0.75, 0.76, 0.77, 0.77, 0.77, 0.78]
    sim.predicted_DER = predicted_DER_List[repeater_delay_multiplier-1]

    sim.networkConfig()

    #Sensor Network
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "ed"):
            env.process(nodes[i].transmit(env))

    #prepare show
    if (sim.graphics == 1):
        sim.plt.xlim([0, sim.xmax])
        sim.plt.ylim([0, sim.ymax])
        sim.plt.draw()
        sim.plt.show()

    # start simulation
    env.run()
    # with open('debug_log_file.txt', 'w') as f:
    #     sys.stdout = f
    #     env.run()

    #-----------------------------------------------------------------------
    #Print simulation stat
    print ("No of nodes: ", len(nodes)) #FIX
    print ("AvgSendTime (exp. distributed):",sim.avgSendTime)
    print ("Experiment: ", sim.experiment)
    print ("Simulation Time: ",env.now/60000,"mins")
    print ("Full Collision: ", sim.full_collision)
    print ("Air time: ", sim.nodes[1].packet[0].rectime)


    # individual packet success rates at each number of hops
    sentSuccessfull = [0]*no_of_repeaters
    sent = [0]*no_of_repeaters
    successRates = [0]*no_of_repeaters
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "ed"):
            # print("Node:",nodes[i].id,"\t-","{:.3f}".format(nodes[i].transmissionSuccessRate()))
            nodes[i].transmissionSuccessRate()
            hops = no_of_repeaters-1-i//(load+1)
            # print(hops)
            sent[hops] += nodes[i].sent
            sentSuccessfull[hops] += nodes[i].sentSuccessful
    for i in range(no_of_repeaters):
        successRates[i] = sentSuccessfull[i]/sent[i]

    print("Sent Pkts:\t",sent)
    print("Successful Sent Pkts:\t",sentSuccessfull)
    print("\nIndividual packet success Rates at each number of hops:")
    for r in successRates:
        print(f"{r:.2f}  ",end="")
    print("")

    # data extraction rate
    der = len(sim.packetsRecBS)/float(sim.totalSimPackets)
    print("\nOverall DER:", der)

    #output rate
    # print("Q1 time:",sim.Q1_time)
    # print("Q2 time:",sim.Q2_time)
    # print("Q3 time:",sim.Q3_time)
    # outputRate = (sim.totalSimPackets*sim.predicted_DER/2)/((sim.Q3_time-sim.Q1_time)/60000)
    # print("Output Rate:",outputRate,"pkts/min")

    #Average Latency
    sum_of_latencies = 0
    sim.packetLatencies.sort()
    for i in range(len(sim.packetLatencies)):
        sum_of_latencies += sim.packetLatencies[i]
    average_latency = sum_of_latencies/len(sim.packetLatencies)

    print("Average Latency:",average_latency,"ms")
    print("Minimum Latency:",sim.packetLatencies[0],"ms")
    print("Maximum Latency:",sim.packetLatencies[-1],"ms")

    # plt.figure(2)
    # plt.hist(sim.packetLatencies, bins=200, edgecolor='black')
    # plt.xlabel('Latency (ms)')
    # plt.ylabel('Number of Packets')

    # Append Test to Excel Sheet
    file_path = "exp2SimData_June2024.xlsx"
    sheet_name = "Sheet"
    values_to_append = []
    values_to_append.append(sim.experiment)
    values_to_append.append(load)
    values_to_append.append(density)
    values_to_append.append(no_of_repeaters)
    values_to_append.append(repeater_delay_multiplier)
    values_to_append.append("---> DER:")
    values_to_append.append(der)
    values_to_append.append("---> Avg Latency:")
    values_to_append.append(average_latency)
    values_to_append.append("---> Min Latency:")
    values_to_append.append(sim.packetLatencies[0])
    values_to_append.append("---> Max Latency:")
    values_to_append.append(sim.packetLatencies[-1])
    # values_to_append.append("---> OUT_Rate:")
    # values_to_append.append(outputRate)
    values_to_append.append("---> Total Collisions:")
    values_to_append.append(len(sim.collidedPackets))
    values_to_append.append("  Fairness:")
    values_to_append +=successRates
    append_values_to_excel(file_path, sheet_name, values_to_append)

    # this can be done to keep graphics visible
    if (sim.graphics == 1):
        input('Press Enter to continue ...')


#============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Repeater Waiting Period Simulation")
    
    # # Add your command-line arguments
    parser.add_argument("-load"                      , default=3 , help="tag load per repeater")
    parser.add_argument("-density"                   , default=2 , help="repeater density")
    parser.add_argument("-no_of_repeaters"           , default=20, help="number of repeaters")
    parser.add_argument("-repeater_delay_multiplier" , default=3 , help="How many times the repeater mean waiting period is greater than the pkt transmission air time?")

    # # Parse the command-line arguments
    args = parser.parse_args()

    # Call the main function with the parsed arguments
    main(int(args.load), int(args.density), int(args.no_of_repeaters), int(args.repeater_delay_multiplier))