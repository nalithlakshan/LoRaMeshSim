import sys
import openpyxl
import argparse
sys.path.append('../../')
import loraMeshSimulator as sim

def append_values_to_excel(file_path, sheet_name, values):
    # Load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the specified sheet or create a new one if it doesn't exist
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    # Append values to a new row
    sheet.append(values)

    # Save the changes to the workbook
    workbook.save(file_path)

def main(load, density):
    #-------------------------------------------------------------------------------------
    # Simulation config
    #-------------------------------------------------------------------------------------
    node = sim.node
    nodes = sim.nodes
    env =sim.env
    maxDist =sim.maxDist

    sim.avgSendTime = 60000 # 5 mins
    sim.repeatDelayMultiplier = 7
    sim.graphics = 1
    sim.realtime_graphics = 0
    sim.debug = 0

    sim.totalSimPackets = 10000

    repeaters =[]
    enddevices = []
    no_of_repeaters = 10
    # density = 1
    # load = 2
    for i in range(no_of_repeaters):
        repeaters.append(node(env, (1+float(i)/density)*maxDist*0.99, 1*maxDist*1, "rp"))
        for j in range(load):
            enddevices.append(node(env, (1+float(i)/density)*maxDist*0.99, 1.99*maxDist, "ed"))

    gw = sim.node(env, (1+float(no_of_repeaters)/density)*maxDist*0.99, 1*maxDist*1, "gw")

    sim.networkConfig()

    #Sensor Network
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "ed"):
            env.process(nodes[i].transmit(env))

    #Actuator Network
    # for i in range(len(nodes)):
    #     if (nodes[i].type.lower() == "gw"):
    #         print("HELLO!!!!!!!!!!!!!")
    #         env.process(nodes[i].transmit(env))

    #prepare show
    if (sim.graphics == 1):
        sim.plt.xlim([0, sim.xmax])
        sim.plt.ylim([0, sim.ymax])
        sim.plt.draw()
        sim.plt.show()

    # start simulation
    env.run()

    #-----------------------------------------------------------------------
    #Print simulation stat
    print ("No of nodes: ", len(nodes)) #FIX
    print ("AvgSendTime (exp. distributed):",sim.avgSendTime)
    print ("Experiment: ", sim.experiment)
    print ("Simulation Time: ",env.now/60000,"mins")
    print ("Full Collision: ", sim.full_collision)
    print ("Air time: ", sim.nodes[1].packet[0].rectime)


    # individual packet success rates at each number of hops
    sentSuccessfull = [0]*no_of_repeaters
    sent = [0]*no_of_repeaters
    successRates = [0]*no_of_repeaters
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "ed"):
            # print("Node:",nodes[i].id,"\t-","{:.3f}".format(nodes[i].transmissionSuccessRate()))
            nodes[i].transmissionSuccessRate()
            hops = no_of_repeaters -i//(load+1)
            sent[hops-1] += nodes[i].sent
            sentSuccessfull[hops-1] += nodes[i].sentSuccessful
    for i in range(no_of_repeaters):
        successRates[i] = sentSuccessfull[i]/sent[i]

    print("Sent Pkts:\t",sent)
    print("Successful Sent Pkts:\t",sentSuccessfull)
    print("\nIndividual packet success Rates at each number of hops (hops= 1 to hops=",no_of_repeaters,"):")
    for r in successRates:
        print(f"{r:.2f}  ",end="")
    print("")


    RpRecPacketCount = [0]*no_of_repeaters
    RpRecPacketPercentage = [0]*no_of_repeaters
    k =0
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "rp"):
            RpRecPacketCount[k] = len(nodes[i].recPackets)
            RpRecPacketPercentage[k] = len(nodes[i].recPackets)/sim.totalSimPackets
            k = k+1
    
    print("\nNo of pkts received by each Rp:")
    for x in RpRecPacketCount:
        print(x," ",end="")
    print("")

    print("\nPercentage pkts received by each Rp:")
    for x in RpRecPacketPercentage:
        print(f"{x:.2f}  ",end="")
    print("")

    # data extraction rate
    der = len(sim.packetsRecBS)/float(sim.totalSimPackets)
    print("\nOverall DER:", der)

    # Append Test to Excel Sheet
    file_path = "repeaterDensitySimData.xlsx"
    sheet_name = "Sheet1"
    values_to_append = []
    values_to_append.append(load)
    values_to_append.append(density)
    values_to_append.append(der)
    values_to_append +=successRates
    append_values_to_excel(file_path, sheet_name, values_to_append)

    # this can be done to keep graphics visible
    if (sim.graphics == 1):
        input('Press Enter to continue ...')


#============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Repeater Density Simulation")
    
    # Add your command-line arguments
    parser.add_argument("arg1", help="tag load per repeater")
    parser.add_argument("arg2", help="repeater density")

    # Parse the command-line arguments
    args = parser.parse_args()

    # Call the main function with the parsed arguments
    main(int(args.arg1), int(args.arg2))